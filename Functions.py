import os
from datetime import datetime
import Main
import openpyxl
from decimal import Decimal


# 判断输入的发票号码是否有效
def validate_invoice_number(invoice_number: str):
    flag1 = invoice_number.isdigit()
    flag2 = len(invoice_number) == 8 or len(invoice_number) == 20

    # print('{}'.format(len(invoice_number)))

    return flag1 and flag2


# 检查发票是否真实
def find_invoice(invoice_number: str):
    # print("获取的发票字符串是： {}".format(invoice_number))

    result = []

    # 获取其中的所有XLSX文件
    file_list = [x for x in os.listdir(Main.DATABASE_PATH) if x.endswith('xlsx')]

    # print(file_list)
    # 遍历发票数据库内所有的Excel文件
    for file_name in file_list:
        full_path = os.path.join(Main.DATABASE_PATH, file_name)
        print(full_path)

        # 打开每一个Excel文件
        wb = openpyxl.load_workbook(full_path)

        # 如果是直接下载的发票数据，仅遍历信息汇总即可
        sheet_names = [x for x in wb.sheetnames]

        # # 遍历其中的所有表，适合小文件
        # sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            # 在每一个表的第三列和第四列寻找相同的字符串
            current_ws = wb[sheet_name]
            print(sheet_name)
            # 只要找到就返回结果
            for i in range(2, current_ws.max_row + 1):
                # 匹配字符串
                if (invoice_number == current_ws.cell(row=i, column=3).value or
                        invoice_number == current_ws.cell(row=i, column=5).value):
                    result.append((invoice_number, current_ws.cell(row=i, column=11).value,
                                   Decimal(current_ws.cell(row=i, column=7).value) +
                                   Decimal(current_ws.cell(row=i, column=8).value),
                                   current_ws.cell(row=i, column=6).value,
                                   file_name))
                # 如果未找到 返回空列表

    return result


# 组装最后的结果
def assemble_find_invoice_result(invoice_list: list, invoice_number: str):
    result = '发票号码：{} 检查结果：\n'.format(invoice_number)
    if len(invoice_list) == 0:
        result += "未找到发票信息，怀疑为假发票"
        return result
    else:
        total_amount = Decimal(0)
        for each_invoice in invoice_list:
            result += '发票号码：{}\t发票抬头：{}\t发票金额：{:.2f}\t开具时间：{}\t文件名称：{}\n'.format(
                each_invoice[0], each_invoice[1], each_invoice[2], each_invoice[3], each_invoice[4])

            total_amount += Decimal(each_invoice[2])

        result += "上述合计金额为：{:.2f}\n".format(total_amount)
        return result


# 到查重台账中寻找发票号码
def find_repeated_invoice(invoice_number: str):
    result = []

    # 根据固定路径打开查重台账
    wb = openpyxl.load_workbook(Main.REPEAT_DATABASE_PATH)
    ws = wb.active

    # 遍历每一行第一列匹配发票号码
    for i in range(1, ws.max_row + 1):
        if invoice_number == ws.cell(row=i, column=1).value == invoice_number:
            result.append((invoice_number, ws.cell(row=i, column=2).value, ws.cell(row=i, column=3).value))

    return result


# 将查到的结果写入到查重台账中，还要记录首次查验的时间
def write_found_invoice_to_repeat_database(invoice_list: list):
    # 打开查重台账
    wb = openpyxl.load_workbook(Main.REPEAT_DATABASE_PATH)

    ws = wb.active

    start_row = ws.max_row + 1

    for each_invoice in invoice_list:
        ws.cell(row=start_row, column=1).value = each_invoice[0]
        ws.cell(row=start_row, column=2).value = each_invoice[1]
        ws.cell(row=start_row, column=3).value = each_invoice[2]
        ws.cell(row=start_row, column=4).value = each_invoice[3]
        ws.cell(row=start_row, column=5).value = datetime.now()
        start_row += 1

    wb.save(Main.REPEAT_DATABASE_PATH)
    wb.close()
